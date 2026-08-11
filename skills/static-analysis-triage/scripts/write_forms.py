#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""判定結果を申請書に書き戻し、逸脱記録書を生成する。

T4（成果物生成）で使う。

    python write_forms.py --clusters work/clusters.json --form-spec work/form-spec.json \
        --out-dir out

原本は絶対に書き換えない。--out-dir にコピーしてから書き、書き込み後に書式検算を行う。
検算に失敗した場合は生成物を破棄して終了コード 1 を返す。

終了コード:
    0  正常。書式検算も通過した
    1  書き込み先の列が無い / 突き合わせ失敗 / 書式検算の失敗（生成物は破棄）
    2  引数の指定ミス、ファイルが無い、openpyxl が無い

書式を壊さないための決まりは references/excel-forms.md を参照。
依存: openpyxl
"""

import argparse
import json
import os
import re
import shutil
import sys

WS_RE = re.compile(r"\s+")

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass


def die(code, msg):
    print("ERROR  %s" % msg, file=sys.stderr)
    sys.exit(code)


def load_openpyxl():
    try:
        import openpyxl
    except ImportError:
        die(2, "openpyxl が無い。pip install openpyxl を実行すること")
    return openpyxl


def load_json(path):
    if not os.path.isfile(path):
        die(2, "見つからない: %s" % path)
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (OSError, ValueError) as e:
        die(2, "読めない: %s (%s)" % (path, e))


def norm_header(value):
    if value is None:
        return ""
    return WS_RE.sub(" ", str(value).replace("　", " ")).strip()


def open_for_write(path, openpyxl):
    """書き込み用に開く。data_only は使わない（数式が値に置き換わって壊れる）。"""
    kwargs = {}
    if path.lower().endswith(".xlsm"):
        kwargs["keep_vba"] = True
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except Exception as e:
        die(2, "ブックを開けない: %s (%s)" % (path, e))


def snapshot(path, openpyxl):
    """検算用のスナップショット。シート名、寸法、全セル値を取る。"""
    wb = openpyxl.load_workbook(path, read_only=True)
    snap = {}
    for name in wb.sheetnames:
        ws = wb[name]
        cells = {}
        for r, row in enumerate(ws.iter_rows(values_only=True), 1):
            for c, v in enumerate(row, 1):
                if v is not None:
                    cells[(r, c)] = str(v)
        snap[name] = {"dim": (ws.max_row or 0, ws.max_column or 0), "cells": cells}
    wb.close()
    return snap


def verify(before, after, sheet, allowed):
    """書き込み対象以外が変化していないことを確かめる。allowed は (行, 列) の集合。"""
    problems = []
    if set(before) != set(after):
        problems.append("シート構成が変わった: %s → %s"
                        % (sorted(before), sorted(after)))
        return problems
    for name in before:
        b, a = before[name], after[name]
        if b["dim"] != a["dim"]:
            problems.append("シート '%s' の寸法が変わった: %s → %s" % (name, b["dim"], a["dim"]))
        keys = set(b["cells"]) | set(a["cells"])
        for key in keys:
            if name == sheet and key in allowed:
                continue
            if b["cells"].get(key) != a["cells"].get(key):
                problems.append("シート '%s' のセル %s が変化した: %r → %r"
                                % (name, key, b["cells"].get(key), a["cells"].get(key)))
                if len(problems) >= 10:
                    problems.append("（以降は省略）")
                    return problems
    return problems


def resolve_write_columns(ws, header_row, columns):
    """書き込み先の列名 → 列番号（1始まり）。存在しない列があれば止める。"""
    header = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        for idx, v in enumerate(row, 1):
            name = norm_header(v)
            if name and name not in header:
                header[name] = idx
        break
    resolved, missing = {}, []
    for logical, col_name in (columns or {}).items():
        if not col_name:
            continue
        key = norm_header(col_name)
        if key in header:
            resolved[logical] = header[key]
        else:
            missing.append("%s → '%s'" % (logical, col_name))
    if missing:
        print("ERROR  書き込み先の列が見つからない: %s" % " / ".join(missing), file=sys.stderr)
        print("       form-spec.json の application_form.columns を確認すること", file=sys.stderr)
        sys.exit(1)
    if not resolved:
        die(1, "書き込み先の列が1つも定義されていない")
    return resolved


def collect_decisions(clusters):
    """行番号 → 判定 の対応を作る。逸脱の一覧も同時に集める。"""
    by_row = {}
    deviations = []

    def put(member, cluster, verdict, code, rationale, alternative):
        row = member.get("row")
        if row is None:
            return
        d = {"verdict": verdict, "code": code or "",
             "rationale": rationale or "", "alternative": alternative or ""}
        by_row[int(row)] = d
        if verdict == "deviate":
            deviations.append({
                "rule_id": cluster.get("rule_id", ""),
                "file": member.get("file", ""),
                "line": member.get("line", ""),
                "message": (cluster.get("representative") or {}).get("message", ""),
                "code": code or "",
                "rationale": rationale or "",
                "alternative": alternative or "",
            })

    for c in clusters:
        members = c.get("members") or []
        if c.get("uniformity") == "individual":
            for m in members:
                put(m, c, m.get("verdict"), m.get("code"),
                    m.get("rationale"), m.get("alternative"))
        else:
            for m in members:
                put(m, c, c.get("verdict"), c.get("code"),
                    c.get("rationale"), c.get("alternative"))
    return by_row, deviations


def write_application(spec, by_row, out_dir, openpyxl):
    src = spec["source"]
    app = spec.get("application_form") or {}
    original = src["workbook"]
    if not os.path.isfile(original):
        die(2, "申請書が見つからない: %s" % original)

    os.makedirs(out_dir, exist_ok=True)
    dest = os.path.join(out_dir, os.path.basename(original))
    shutil.copy2(original, dest)

    before = snapshot(original, openpyxl)
    wb = open_for_write(dest, openpyxl)
    sheet = src.get("sheet")
    if sheet not in wb.sheetnames:
        os.remove(dest)
        die(1, "シート '%s' が無い" % sheet)
    ws = wb[sheet]

    header_row = int(src.get("header_row") or 1)
    cols = resolve_write_columns(ws, header_row, app.get("columns"))

    # 書き込み先が指摘データの列と重なっていないか確かめる。重なった場合、そのセルは
    # 検算の対象外（書き込み対象なので当然変化する）になり、上書きを検出できない。
    src_names = set()
    for name in (src.get("columns") or {}).values():
        if name:
            src_names.add(norm_header(name))
    if src.get("id_column"):
        src_names.add(norm_header(src["id_column"]))
    src_cols = resolve_write_columns(ws, header_row, dict(
        ("src_%d" % i, n) for i, n in enumerate(sorted(src_names))))
    overlap = sorted(set(cols.values()) & set(src_cols.values()))
    if overlap:
        wb.close()
        os.remove(dest)
        inv = dict((v, k) for k, v in cols.items())
        print("ERROR  書き込み先の列が指摘データの列と重なっている: %s"
              % ", ".join("%d列目(%s)" % (c, inv.get(c, "?")) for c in overlap), file=sys.stderr)
        print("       このまま書くと元の指摘が上書きされ、検算でも検出できない。"
              "form-spec.json の application_form.columns を別の列に直すこと", file=sys.stderr)
        sys.exit(1)

    labels = app.get("verdict_labels") or {}

    written = 0
    touched = set()
    for row, d in sorted(by_row.items()):
        if row <= header_row or row > (ws.max_row or 0):
            continue
        values = {
            "verdict": labels.get(d["verdict"], d["verdict"]),
            "code": d["code"],
            "rationale": d["rationale"],
            "alternative": d["alternative"],
        }
        for logical, col in cols.items():
            ws.cell(row=row, column=col).value = values.get(logical, "")
            touched.add((row, col))
        written += 1

    wb.save(dest)
    wb.close()

    problems = verify(before, snapshot(dest, openpyxl), sheet, touched)
    if problems:
        os.remove(dest)
        print("ERROR  書式検算に失敗した。生成物は破棄した", file=sys.stderr)
        for p in problems:
            print("       %s" % p, file=sys.stderr)
        print("       検算を無効化して押し通さないこと。"
              "form-spec.json の書き込み列を確認すること", file=sys.stderr)
        sys.exit(1)

    unmatched = len(by_row) - written
    print("申請書: %s" % dest)
    print("  書き込み %d 行 / 判定 %d 件（対象外の行 %d）" % (written, len(by_row), unmatched))
    print("  書き込んだ列: %s" % ", ".join(sorted(cols)))
    print("  書式検算: 合格（シート構成・寸法・対象外セルに変化なし）")
    if written == 0:
        die(1, "1行も書き込めなかった。clusters.json の members[].row と申請書の行が対応していない")
    return dest


def write_deviation_record(spec, deviations, out_dir, openpyxl):
    rec = spec.get("deviation_record") or {}
    cols = rec.get("columns") or {}
    if not cols:
        print("逸脱記録書: form-spec.json に deviation_record.columns が無いため生成しない")
        return None

    template = rec.get("template")
    sheet = rec.get("sheet") or "逸脱記録"
    header_row = int(rec.get("header_row") or 1)
    start_row = int(rec.get("start_row") or (header_row + 1))

    dest = os.path.join(out_dir, "逸脱記録書.xlsx")
    if template:
        if not os.path.isfile(template):
            die(2, "逸脱記録書の雛形が見つからない: %s" % template)
        dest = os.path.join(out_dir, os.path.basename(template))
        shutil.copy2(template, dest)
        before = snapshot(template, openpyxl)
        wb = open_for_write(dest, openpyxl)
        if sheet not in wb.sheetnames:
            os.remove(dest)
            die(1, "逸脱記録書の雛形にシート '%s' が無い" % sheet)
        ws = wb[sheet]
        col_index = resolve_write_columns(ws, header_row, cols)
    else:
        # 雛形が無い場合は新規ブックを作る。社内様式があるなら template を指定すること。
        before = None
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        order = [k for k in ("seq", "rule_id", "file", "line", "code",
                             "rationale", "alternative", "message") if cols.get(k)]
        col_index = {}
        for i, logical in enumerate(order, 1):
            ws.cell(row=header_row, column=i).value = cols[logical]
            col_index[logical] = i

    touched = set()
    for i, dev in enumerate(deviations):
        row = start_row + i
        values = dict(dev)
        values["seq"] = i + 1
        for logical, col in col_index.items():
            ws.cell(row=row, column=col).value = values.get(logical, "")
            touched.add((row, col))

    wb.save(dest)
    wb.close()

    if before is not None:
        problems = verify(before, snapshot(dest, openpyxl), sheet, touched)
        # 雛形は空行を持つため、行数の増加は正当な変化として許容する。
        problems = [p for p in problems if "寸法が変わった" not in p]
        if problems:
            os.remove(dest)
            print("ERROR  逸脱記録書の書式検算に失敗した。生成物は破棄した", file=sys.stderr)
            for p in problems:
                print("       %s" % p, file=sys.stderr)
            sys.exit(1)

    print("逸脱記録書: %s" % dest)
    print("  書き込み %d 行（%d 行目から）" % (len(deviations), start_row))
    if not template:
        print("  ※ 雛形が指定されていないため新規ブックとして生成した。"
              "社内様式がある場合は form-spec.json の deviation_record.template に指定すること")
    return dest


def main(argv):
    ap = argparse.ArgumentParser(description="判定結果の申請書への書き戻しと逸脱記録書の生成")
    ap.add_argument("--clusters", required=True, metavar="JSON", help="work/clusters.json")
    ap.add_argument("--form-spec", required=True, metavar="JSON", help="work/form-spec.json")
    ap.add_argument("--out-dir", default="out", metavar="DIR", help="出力先（既定: out）")
    args = ap.parse_args(argv)

    openpyxl = load_openpyxl()
    spec = load_json(args.form_spec)
    if "source" not in spec:
        die(1, "form-spec に source ブロックが無い")

    data = load_json(args.clusters)
    clusters = data.get("clusters") or []
    if not clusters:
        die(1, "クラスタが0件: %s" % args.clusters)

    by_row, deviations = collect_decisions(clusters)
    if not by_row:
        die(1, "判定が1件も無い。先に T3 を終えること")

    write_application(spec, by_row, args.out_dir, openpyxl)
    if deviations:
        write_deviation_record(spec, deviations, args.out_dir, openpyxl)
    else:
        print("逸脱記録書: deviate が0件のため生成しない")

    deferred = sum(1 for d in by_row.values() if d["verdict"] == "deferred")
    if deferred:
        print()
        print("WARN   deferred が %d 件残っている。成果物にその旨を明記し、"
              "伏せて提出しないこと" % deferred)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
