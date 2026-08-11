#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""申請書 Excel から静的解析の指摘一覧を読み出す。

T0（様式の確定）と T1（取り込み）で使う。

    python read_findings.py --inspect <ブック.xlsx>
    python read_findings.py --form-spec work/form-spec.json -o work/findings.json

--inspect はセルの中身をほとんど返さない。シート名、行数・列数、ヘッダ行の候補と
その列名だけを出す。数千行の申請書を会話に載せないための入口である。

終了コード:
    0  正常
    1  様式の不整合（シートが無い / 必須列が無い / 有効な行が0件）
    2  引数の指定ミス、ファイルが無い、openpyxl が無い

form-spec.json の書き方は references/excel-forms.md を参照。
依存: openpyxl
"""

import argparse
import json
import os
import re
import sys

REQUIRED_COLUMNS = ("rule_id", "file", "message")
OPTIONAL_COLUMNS = ("line", "tool", "severity", "function", "code")
INSPECT_SCAN_ROWS = 10       # ヘッダ行の候補を探す範囲（先頭から何行か）
INSPECT_SAMPLE_CHARS = 40    # 列名として表示する最大文字数

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass


def die(code, msg):
    print("ERROR  %s" % msg, file=sys.stderr)
    sys.exit(code)


def load_openpyxl():
    try:
        import openpyxl
    except ImportError:
        die(2, "openpyxl が無い。pip install openpyxl を実行すること")
    return openpyxl


def norm_header(value):
    """列名を突き合わせ用に正規化する。全角空白・改行・連続空白を畳む。"""
    if value is None:
        return ""
    s = str(value).replace("　", " ")
    return re.sub(r"\s+", " ", s).strip()


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def open_book(path, openpyxl):
    if not os.path.isfile(path):
        die(2, "ブックが見つからない: %s" % path)
    try:
        # data_only=True で数式の代わりに直近の計算結果を読む。
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        die(2, "ブックを開けない: %s (%s)" % (path, e))


def inspect(path):
    openpyxl = load_openpyxl()
    wb = open_book(path, openpyxl)
    print("ブック: %s" % path)
    print("シート数: %d" % len(wb.sheetnames))
    for name in wb.sheetnames:
        ws = wb[name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        print()
        print("== シート '%s'  行数 %d / 列数 %d" % (name, rows, cols))
        if rows == 0 or cols == 0:
            print("   （空）")
            continue

        # 先頭 INSPECT_SCAN_ROWS 行を読み、非空セルが最も多い行をヘッダ行の候補とする。
        head = []
        for n, row in enumerate(ws.iter_rows(min_row=1, max_row=min(rows, INSPECT_SCAN_ROWS),
                                             values_only=True), 1):
            head.append((n, [norm_header(v) for v in row]))
        if not head:
            print("   （空）")
            continue
        best_n, best_cells = max(head, key=lambda t: sum(1 for c in t[1] if c))
        filled = sum(1 for c in best_cells if c)
        print("   ヘッダ行の候補: %d 行目（非空セル %d 個）" % (best_n, filled))
        print("   ※ 候補は推定にすぎない。タイトル行が上にある様式では外すので目で確認すること")
        for idx, c in enumerate(best_cells, 1):
            if c:
                shown = c if len(c) <= INSPECT_SAMPLE_CHARS else c[:INSPECT_SAMPLE_CHARS] + "…"
                print("     列%-3d %s" % (idx, shown))
    wb.close()
    print()
    print("この出力の列名をそのまま form-spec.json にコピーすること（手で打ち直さない）。")
    return 0


def load_spec(path):
    if not os.path.isfile(path):
        die(2, "form-spec が見つからない: %s" % path)
    try:
        with open(path, "r", encoding="utf-8") as f:
            spec = json.load(f)
    except (OSError, ValueError) as e:
        die(2, "form-spec を読めない: %s (%s)" % (path, e))
    if "source" not in spec:
        die(1, "form-spec に source ブロックが無い")
    return spec


def resolve_columns(header_cells, spec_columns, sheet_name):
    """列名 → 列インデックス（0始まり）に解決する。"""
    index = {}
    for idx, name in enumerate(header_cells):
        if name and name not in index:
            index[name] = idx

    resolved = {}
    missing = []
    for logical in REQUIRED_COLUMNS + OPTIONAL_COLUMNS:
        wanted = spec_columns.get(logical)
        if not wanted:
            continue
        key = norm_header(wanted)
        if key in index:
            resolved[logical] = index[key]
        elif logical in REQUIRED_COLUMNS:
            missing.append("%s → '%s'" % (logical, wanted))
        else:
            print("WARN   任意列 %s の列名 '%s' がシート '%s' に無い。この項目は使わない"
                  % (logical, wanted, sheet_name))
    if missing:
        print("ERROR  必須列が見つからない: %s" % " / ".join(missing), file=sys.stderr)
        print("       ヘッダ行の指定がずれていないか、--inspect の出力と突き合わせること",
              file=sys.stderr)
        sys.exit(1)
    return resolved


def extract(spec, out_path):
    openpyxl = load_openpyxl()
    src = spec["source"]
    wb = open_book(src["workbook"], openpyxl)

    sheet_name = src.get("sheet")
    if sheet_name not in wb.sheetnames:
        print("ERROR  シート '%s' が無い。存在するのは: %s"
              % (sheet_name, ", ".join(wb.sheetnames)), file=sys.stderr)
        sys.exit(1)
    ws = wb[sheet_name]

    header_row = int(src.get("header_row") or 1)
    header_cells = []
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        header_cells = [norm_header(v) for v in row]
        break
    if not any(header_cells):
        print("ERROR  %d 行目が空。header_row の指定を確認すること" % header_row, file=sys.stderr)
        sys.exit(1)

    cols = resolve_columns(header_cells, src.get("columns") or {}, sheet_name)
    id_col_name = norm_header(src.get("id_column") or "")
    id_col = header_cells.index(id_col_name) if id_col_name in header_cells else None

    findings = []
    dropped = {"empty": 0, "header_repeat": 0, "missing_required": 0}
    header_set = set(c for c in header_cells if c)

    for n, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                            header_row + 1):
        cells = [cell_text(v) for v in row]
        if not any(cells):
            dropped["empty"] += 1
            continue
        # ページ区切りなどでヘッダが再掲される様式がある。
        if header_set and header_set.issubset(set(c for c in cells if c)):
            dropped["header_repeat"] += 1
            continue

        rec = {}
        for logical, idx in cols.items():
            rec[logical] = cells[idx] if idx < len(cells) else ""
        if not all(rec.get(k) for k in REQUIRED_COLUMNS):
            dropped["missing_required"] += 1
            continue

        rec["row"] = n
        rec["id"] = (cells[id_col] if id_col is not None and id_col < len(cells) else "")
        findings.append(rec)

    wb.close()

    total_scanned = len(findings) + sum(dropped.values())
    print("シート '%s' の %d 行目以降を走査した" % (sheet_name, header_row + 1))
    print("取り込み: %d 件" % len(findings))
    print("除外: 空行 %d / ヘッダ再掲 %d / 必須列が空 %d"
          % (dropped["empty"], dropped["header_repeat"], dropped["missing_required"]))

    if not findings:
        print("ERROR  有効な行が0件。header_row か columns の指定が誤っている可能性が高い",
              file=sys.stderr)
        sys.exit(1)

    # 空行はページ末尾の余白で普通に出るので、判断材料からは外す。
    meaningful_dropped = dropped["header_repeat"] + dropped["missing_required"]
    if total_scanned and meaningful_dropped > total_scanned * 0.05:
        print("WARN   空行以外の除外が %d 件（走査 %d 行の 5%% 超）。"
              "様式の読み違いを疑い、T0 に戻って列対応を確認すること"
              % (meaningful_dropped, total_scanned))

    payload = {
        "version": 1,
        "source_workbook": src["workbook"],
        "source_sheet": sheet_name,
        "header_row": header_row,
        "columns_used": sorted(cols.keys()),
        "findings": findings,
    }
    out_dir = os.path.dirname(os.path.abspath(out_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print("出力: %s" % out_path)
    return 0


def main(argv):
    ap = argparse.ArgumentParser(description="申請書 Excel から静的解析の指摘一覧を読み出す")
    ap.add_argument("--inspect", metavar="XLSX",
                    help="ブックの構造だけを出す（セルの中身は返さない）")
    ap.add_argument("--form-spec", metavar="JSON", help="work/form-spec.json")
    ap.add_argument("-o", "--out", metavar="JSON", default="work/findings.json",
                    help="出力先（既定: work/findings.json）")
    args = ap.parse_args(argv)

    if args.inspect:
        return inspect(args.inspect)
    if not args.form_spec:
        ap.error("--inspect か --form-spec のどちらかを指定すること")
    return extract(load_spec(args.form_spec), args.out)


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
